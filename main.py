import os
import shutil
from openpyxl import load_workbook
from PIL import Image
from PIL import ImageDraw
from PIL import ImageFont
import qrcode



src_dir = os.getcwd() #get the current working dir
def copyAndRename(badge):
    # print(src_dir)

    # create a dir where we want to copy and rename
    # dest_dir = os.mkdir('Badges')
    # os.listdir()

    dest_dir = src_dir+"/Badges"
    src_file = os.path.join(src_dir, badge)
    shutil.copy(src_file,dest_dir) #copy the file to destination dir

    dst_file = os.path.join(dest_dir,badge)
    new_dst_file_name = os.path.join(dest_dir, 'name second name.png')

    os.rename(dst_file, new_dst_file_name)#rename
    os.chdir(dest_dir)

    return new_dst_file_name

def break_fix(text, width, font, draw):
    if not text:
        return
    lo = 0
    hi = len(text)
    while lo < hi:
        mid = (lo + hi + 1) // 2
        t = text[:mid]
        w, h = draw.textsize(t, font=font)
        if w <= width:
            lo = mid
        else:
            hi = mid - 1
    t = text[:lo]
    w, h = draw.textsize(t, font=font)
    yield t, w, h
    yield from break_fix(text[lo:], width, font, draw)

def fit_text(img, text, color, font):
    width = img.size[0] - 2
    draw = ImageDraw.Draw(img)
    pieces = list(break_fix(text, width, font, draw))
    height = sum(p[2] for p in pieces)
    if height > img.size[1]:
        raise ValueError("text doesn't fit")
    y = (img.size[1] - height) // 2
    for t, w, h in pieces:
        x = (img.size[0] - w) // 2
        draw.text((x, y), t, font=font, fill=color)
        y += h


def writeTxtOnBadge(name, affiliation,aff2, badge_img,name2="", aff3 = ""):
    # file = copyAndRename()
    
    im = Image.open(badge_img)
    W, H = im.size

    # x = W/10
    # y1 = H/2.5

    # Call draw Method to add 2D graphics in an image
    I1 = ImageDraw.Draw(im)
     # Custom font style and font size
    myFont = ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 72)
    myFont2 = ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 55)
     
    # Add Text to an image
    # I1.text((x, y), "Nice Car", font=myFont, fill =(255, 255, 255))
    _, _, w, h = I1.textbbox((20, 0), name, font=myFont)
    if w >= W + 10:
        myFont =  ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 62)
        _, _, w, h = I1.textbbox((0, 0), name, font=myFont)

    I1.text(((W-w)/2, (H-h)/2.3), name, font=myFont, fill=(255,255,255))

    _, _, wn, hn = I1.textbbox((20, 0), name2, font=myFont)
    if wn >= W + 10:
        myFont =  ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 62)
        _, _, wn, hn = I1.textbbox((0, 0), name2, font=myFont)

    I1.text(((W-wn)/2, (H-hn)/2.05), name2, font=myFont, fill=(255,255,255))

    
    _, _, w2, h2 = I1.textbbox((20, 0), affiliation, font=myFont2)
    if w2 >= W + 10:
        myFont2 =  ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 45)
        _, _, w2, h2 = I1.textbbox((20, 0), affiliation, font=myFont2)

    I1.text(((W-w2)/2, (H-h2)/1.85), affiliation, font=myFont2, fill=(255,255,255))

    _, _, w3, h3 = I1.textbbox((20, 0), aff2, font=myFont2)
    if w3 >= W + 10:
        myFont2 =  ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 45)
        _, _, w3, h3 = I1.textbbox((20, 0), aff2, font=myFont2)

    I1.text(((W-w3)/2, (H-h3)/1.70), aff2, font=myFont2, fill=(255,255,255))

    _, _, w4, h4 = I1.textbbox((20, 0), aff3, font=myFont2)
    if w4 >= W + 10:
        myFont2 =  ImageFont.truetype(src_dir + '/montesrrat-bold.ttf', 45)
        _, _, w4, h4 = I1.textbbox((20, 0), aff3, font=myFont2)

    I1.text(((W-w4)/2, (H-h4)/1.52), aff3, font=myFont2, fill=(255,255,255))


    # Display edited image
    # im.show()
     
    # Save the edited image
    badgePath = src_dir+ "/Badges/" + name + '.png'
    im.save(badgePath)
    return badgePath

def generateQr(data:str, size=13):
    # Creating an instance of QRCode class
    qr = qrcode.QRCode(version = 1,
                       box_size = size,
                       border = 0)
    
    # Adding data to the instance 'qr'
    qr.add_data(data)
    
    qr.make(fit = True)
    img = qr.make_image(fill_color = '#1c2813',
    # img = qr.make_image(fill_color = 'black',
                        back_color = 'white')
    
    qrPath = src_dir + '/Badges/QR_' +data+'.png'
    img.save(qrPath)

    return qrPath

def putQrOnBadge(qrPath, badgePath):
    # Opening the primary image (used in background) 
    img1 = Image.open(badgePath) 
    W, H = img1.size
      
    # Opening the secondary image (overlay image) 
    img2 = Image.open(qrPath) 
    w, h = img2.size 
    # Pasting img2 image on top of img1  
    # starting at coordinates (0, 0) 
    w1 = w  + w/7
    h1 = h + w/10
    img1.paste(img2, (int(W-w1),int(H-h1)))
      
    # Displaying the image 
    # img1.show()

    img1.save(badgePath)

wb = load_workbook('Badges_Names.xlsx')
ws = wb["Names in Badge"]

# Name in Badge (B)
# Affiliation (C)
# Role (D)
def main():
    for i in range(2,127):
        name = ws.cell(row = i, column=2).value
        aff = ws.cell(row = i, column=3).value
        aff2 = ws.cell(row = i, column=4).value
        role = ws.cell(row=i, column=5).value 

        if aff2 == None:
            aff2 = ""
        print(name)

        back = src_dir + "/" + str(role) + ".png"
        badge_path = writeTxtOnBadge(name, aff,aff2,back)
        qr_path = generateQr(str(name))
        putQrOnBadge(qr_path, badge_path)
        os.remove(qr_path)

def custom():
    name = "Ceslause Okechukwu "
    name2 = "Ogbonnaya"
    aff = "" 
    aff2 = "Igbo Wikimedians UG" 
    aff3 = "Wiki In Africa"
    role = "Speaker"

    print(name)

    back = src_dir + "/" + str(role) + ".png"
    badge_path = writeTxtOnBadge(name, aff,aff2,back,name2, aff3)
    qr_path = generateQr(name+name2, 12)
    putQrOnBadge(qr_path, badge_path)
    os.remove(qr_path)

# main()
custom()
